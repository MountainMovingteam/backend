import json

import openpyxl
import xlrd

from mysite.lib import time
from .models import Team
from mysite.lib.static_fun import *


# Create your views here.

def person(request):
    token = request.META.get('HTTP_AUTHORIZATION')
    if not token:
        return none_token()
    id, role, is_login = check_token(token)
    if not is_login:
        return login_timeout()
    user = get_user(id, role)
    if user is None:
        return user_not_exists()

    data = json.loads(request.body.decode('utf-8'))

    week_num, time_index = time.trans_index(data['time_index'])
    if week_num < 0 or week_num > WEEK_NUM:
        return week_num_illegal()

    if time_index < 0 or time_index > TIMR_INDEX_NUM:
        return time_index_illegal()

    place = Place.objects.filter(week_num=week_num, time_index=time_index).first()
    if place is None:
        return place_not_exists()

    if place.capacity > count_order(place.id):
        Order.objects.create(place=place, user_id=id, name=data['name'], phone=data['phone'],
                             academy=data['academy'], is_person=True)
        return success_respond()
    else:
        return place_full()


def group(request):
    token = request.META.get('HTTP_AUTHORIZATION')
    if not token:
        return none_token()
    id, role, is_login = check_token(token)
    if not is_login:
        return login_timeout()
    user = get_user(id, role)
    if user is None:
        return user_not_exists()

    data = json.loads(request.body.decode('utf-8'))
    week_num, time_index = time.trans_index(data['time_index'])
    if week_num < 0 or week_num > WEEK_NUM:
        return week_num_illegal()

    if time_index < 0 or time_index > TIMR_INDEX_NUM:
        return time_index_illegal()

    place = Place.objects.filter(week_num=week_num, time_index=time_index).first()
    if place is None:
        return place_not_exists()

    if place.capacity <= count_order(place.id):
        return place_full()

    if have_group(place.id):
        return place_has_group()

    team = Team.objects.create(leader_name=data['leader']['name'], leader_id=id,
                               leader_phone=data['leader']['phone'], academy=data['academy'])
    persons = data['persons']
    for person in persons:
        TeamMember.objects.create(team=team, member_id=person['id'], member_name=person['name'])
    Order.objects.create(place=place, user_id=id, name=data['leader']['name'],
                         phone=data['leader']['phone'],
                         academy=data['academy'], is_person=False, team=team)
    return success_respond()


def get_info(request):
    response = user_auth(request)

    if response is not None:
        return response

    details = []
    for time_index in range(1, 57, 1):
        week_num, new_time_index = time.trans_index(time_index)
        place = Place.objects.filter(week_num=week_num, time_index=new_time_index).first()
        if place is None:
            continue
        capacity = 20
        enrolled = count_order(place.id)
        lecturer = ""
        lecturer_with_place = place.lecturerplace_set.all()
        for l in lecturer_with_place:
            lecturer += " " + Lecturer.objects.filter(id=l.lecturer_id).first().name
        if enrolled == 0:
            type = 0
        elif capacity == enrolled:
            type = 2
        else:
            type = 1
        de = {
            "time_index": time_index,
            "capacity": capacity,
            "enrolled": enrolled,
            "lecturer": lecturer,
            "type": type
        }
        details.append(de)
    return JsonResponse(details, safe=False)


def delete_order(request):
    token = request.META.get('HTTP_AUTHORIZATION')
    if not token:
        return none_token()
    id, role, is_login = check_token(token)
    if not is_login:
        return login_timeout()
    user = get_user(id, role)
    if user is None:
        return user_not_exists()

    data = json.loads(request.body.decode('utf-8'))
    week_num = data['week_num']
    time_index = data['time_index']

    place = Place.objects.filter(week_num=week_num, time_index=time_index).first()
    if place is None:
        return place_not_exists()

    order = Order.objects.filter(user_id=id, place=place).first()
    if order is None:
        return order_not_exists()

    if current_time_index() >= time_index2fabs(week_num, time_index):
        is_expired = True
    else:
        is_expired = False

    if is_expired:
        return order_has_expired()

    order.delete()
    return success_respond()


def init_place(request):
    response = admin_auth(request)
    if response is not None:
        return response
    for week_num in range(1, WEEK_NUM + 1, 1):
        for time_index in range(1, TIMR_INDEX_NUM + 1, 1):
            if len(Place.objects.filter(week_num=week_num, time_index=time_index)) == 0:
                Place.objects.create(week_num=week_num, time_index=time_index, capacity=PLACE_CAPACITY)
    return JsonResponse({'success': True})


def clear_place(request):
    response = admin_auth(request)
    if response is not None:
        return response

    Place.objects.all().delete()
    return JsonResponse({'success': True})


def count_order(place_id):
    count = Order.objects.filter(place_id=place_id, is_person=True).count()
    order_with_team = Order.objects.filter(place_id=place_id, is_person=False).select_related('team').first()
    if order_with_team:
        team_with_member = TeamMember.objects.select_related('team').filter(team_id=order_with_team.team_id).all()
        count += len(team_with_member)
    return count


def have_group(place_id):
    count = Order.objects.filter(place_id=place_id, is_person=False).count()
    return count == 1


def paser_excel(request):
    response = user_auth(request)

    if response is not None:
        return response

    result = []

    if request.FILES.get('file', None):
        excel = request.FILES['file']
        wb = openpyxl.load_workbook(excel)
        print(wb)
        ws = wb.active
        print(ws)
        if ws.cell(row=1, column=1).value == "学号" and ws.cell(row=1, column=2).value == "姓名":
            for row in ws.iter_rows(min_row=2):
                r1 = row[1].value
                r0 = row[0].value
                print(r1, r0)
                print(type(r1),type(r0))
                if (r1 is None or r1 == "") and (r0 is None or r0 == ""):
                    continue
                else:
                    name = id = ''
                    if r1:
                        name = str(r1)
                    if r0:
                        id = str(r0)
                    person = {
                        'name': name,
                        'id': id
                    }
                    result.append(person)
        else:
            return excel_error()
        return JsonResponse({
            "list": result
        })
    else:
        print("none file")
        return excel_error()
